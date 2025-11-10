# excel_postprocessor.py
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.pagebreak import PageBreak, Break


class DailyReportPostProcessor:
    TARGET_TEXT = "花卉(把)"          # 第2頁第一列
    SEARCH_MAX_COL = 21               # A..U
    SEARCH_START_ROW = 1
    DATA_ROW_START = 9
    DATA_ROW_END_FALLBACK = 152       # 覆蓋到說明區
    COL_M, COL_S = 13, 19             # M~S

    # def __init__(self, file_path: str):
    #     self.file_path = str(file_path)

    def _hide_rows_with_dash_in_M_to_S(self, ws, start_row: int, end_row: int, keep_rows=None):        
        keep = set(keep_rows or [])  # 例如 {119, 135} : 保留豬 -> 118
        end_row = min(end_row, ws.max_row or end_row)
        for r in range(start_row, end_row + 1):
            if r in keep:
                continue
            vals = []
            for c in range(self.COL_M, self.COL_S + 1):
                v = ws.cell(row=r, column=c).value
                vals.append(str(v).strip() if v is not None else "")
            if vals and all(v == "-" for v in vals):
                ws.row_dimensions[r].hidden = True


    def _find_row_contains(self, ws, text: str, min_row: int, max_row: int, max_col: int):
        target = str(text)
        for r in range(min_row, max_row + 1):
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and target in str(v).strip():
                    return r
        return None


    def _last_filled_row_in_AU(self, ws) -> int:
        last = 0
        max_r = ws.max_row or 0
        for r in range(1, max_r + 1):
            for c in range(1, self.SEARCH_MAX_COL + 1):  # A..U
                v = ws.cell(row=r, column=c).value
                if v not in (None, ""):
                    last = max(last, r)
        return last

    def process(self, wb: Workbook):
        ws = wb[wb.sheetnames[0]]
        self._hide_rows_with_dash_in_M_to_S(ws, self.DATA_ROW_START, self.DATA_ROW_END_FALLBACK, keep_rows= [118])

        last_raw = self._last_filled_row_in_AU(ws)
        target_raw = self._find_row_contains(ws, self.TARGET_TEXT, self.SEARCH_START_ROW, last_raw, self.SEARCH_MAX_COL)

        ws.print_area = f"A1:U{last_raw}"


        # Characteristic in old version's openpyxl: we need to update page_breaks attribute to update break points in page
        # Note: Basic kownledge of memory address allocation in tuple: if we need to update the values in tuple, 
        # we should assign not only elements, but also whole tuple variable 
        ws.row_breaks = PageBreak()
        ws.col_breaks  = PageBreak() 
        ws.row_breaks.append(Break(id=target_raw - 1))  # Update break points in row_breaks

        ws.page_breaks = (ws.row_breaks, ws.col_breaks)  # We need to assign new elements' values into tuple variable (Recap pointer, references knowledge.)
        
        return wb


        